import matplotlib.pyplot as plt
from matplotlib import gridspec
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl as exl
import decode

columns = decode.data_types

def get_cell_value (sheet, col, row):
    cell_name = "{}{}".format(col, row)
    return sheet[cell_name].value

def set_cell_value (sheet, col, row, value):
    sheet.cell(row = row, column = col).value = value

def open_read_write_new_csvfile(file1, file2):
    '''Reads from given csv file, and extracts given columns from given rows and writes to another given csv file.'''

    with open(file1) as fin:
        fout = open(file2, 'a')
        i = 0
        for line in fin:
            try:
                line = int(line.strip()) #strip whitespace

                #decode the data
                decoded = decode.decode_data(line)

                #create new string to represent line
                new_ln = ', '.join(str(e) for e in decoded) + '\n'

                fout.write(new_ln)
                
            except:
                print(i)
            finally:
                i += 1
        fout.close()

def find_problem_rows(data_file):
    with open(data_file) as fin:
        i = 0
        for line in fin:
            try:
                line = int(line.strip()) #strip whitespace

                #decode the data
                decoded = decode.decode_data(line)
                
            except:
                print(i)
            finally:
                i += 1

def fill_in_time_gaps(sheet):
    #fill in the time gaps
    time_col = []
    #here you iterate over the rows in the specific column
    for row in range(2,sheet.max_row+1):  
        time_col.append(get_cell_value(sheet, "A", row))

    #debug feature
    # for time in time_col:
    #     if not (time == None or np.isnan(time)):
    #         print(time)

    #now go back through and fill in the time gaps
    df = pd.DataFrame(time_col)
    df.interpolate(inplace=True, mode="linear", limit_area="inside")

    i = 0
    col = columns.index("time_stamp")+1
    for row in range(2, sheet.max_row+1):
        set_cell_value(sheet, col, row, df.iat[i, 0])
        i += 1

def fill_in_bio_gaps(sheet):
    #fill in the spo2 gaps
    spo2 = []
    #here you iterate over the rows in the specific column
    for row in range(2,sheet.max_row+1):  
        spo2.append(get_cell_value(sheet, "I", row))

    #now go back through and fill in the spo2 gaps
    df = pd.DataFrame(spo2)
    df.fillna(method='pad', inplace=True)

    i = 0
    col = columns.index("spo2")+1
    for row in range(2, sheet.max_row+1):
        set_cell_value(sheet, col, row, df.iat[i, 0])
        i += 1

    #fill in the hr gaps
    hr = []
    #here you iterate over the rows in the specific column
    for row in range(2,sheet.max_row+1):  
        hr.append(get_cell_value(sheet, "J", row))

    #now go back through and fill in the hr gaps
    df = pd.DataFrame(hr)
    df.fillna(method='pad', inplace=True)

    i = 0
    col = columns.index("hr")+1
    for row in range(2, sheet.max_row+1):
        set_cell_value(sheet, col, row, df.iat[i, 0])
        i += 1
    


def write_to_sheet(txt, sheet):    
    '''Reads from given txt file, and sorts data to excel file'''
    
    row = dict.fromkeys(columns)
    sheet.append(columns)

    with open(txt) as fin:
        i = 0
        for line in fin:
            try:
                line = int(line.strip()) #strip whitespace

                #decode the data
                decoded = decode.decode_data(line)
                value = decoded[0]
                type = decoded[1]
                index = decoded[2]

                #see if we need a new row
                if( row[type] != None ):
                    values = list(row.values())
                    sheet.append(values)
                    row = dict.fromkeys(columns)

                #add data
                row[type] = value
                
            except:
                print("problem",i)
            finally:
                i += 1
    fill_in_time_gaps(sheet)
    # fill_in_bio_gaps(sheet) 

def current_time():
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    return current_time

def create_sheet(wb, data_file, sheet_name):
    wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]
    print("Sheet created:", current_time())

    #pass the data to the excel sheet
    write_to_sheet(data_file, sheet)
    print("Sheet written to excel:", current_time())


def main():
    
    start_time = current_time()
    print("Start Time:", start_time)

    #be sure to CD into the right directory
    wb = exl.Workbook()

    # ####Test
    data_file = "YOUR_FILE_PATH1"
    create_sheet(wb, data_file, "DELAY")

    # ####Test
    data_file = "YOUR_FILE_PATH2"
    create_sheet(wb, data_file, "NO_DELAY")

    #save the file
    wb.save("YOUR_FILE_PATH")

    # data_file = "YOUR_FILE_PATH"
    # find_problem_rows(data_file)

    print('Done:', current_time())

if __name__ == '__main__':
    main()